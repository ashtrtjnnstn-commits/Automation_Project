from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.models import (
    AdminAttendance,
    AdminStaff,
    AttendanceSession,
    BillingAdvice,
    Payment,
    PaymentAllocation,
    RegularSchedule,
    SessionOverride,
    Student,
    Therapist,
    db,
)
from app.services.attendance_service import weekly_therapist_hours

DAY_MAP = {"monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3, "friday": 4, "saturday": 5, "sunday": 6}


def _end_time_from_duration(start_time, duration_hours: float):
    dt = datetime.combine(date.today(), start_time) + timedelta(hours=duration_hours)
    return dt.time().replace(second=0, microsecond=0)


def _sheet_header(ws) -> list[str]:
    return [str(c.value).strip() if c.value else "" for c in ws[1]]


def _validate_columns(header: list[str], mapped: dict[str, str]) -> dict[str, int]:
    missing = [sheet_col for sheet_col in mapped.values() if sheet_col not in header]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return {k: header.index(v) for k, v in mapped.items()}


def import_students_and_schedules(path: str, column_map: dict[str, str] | None = None) -> int:
    """Import students and regular schedules from a single worksheet."""
    wb = load_workbook(path)
    ws = wb.active
    header = _sheet_header(ws)
    mapped = column_map or {
        "student_name": "Student Name",
        "therapist_name": "Therapist",
        "day_of_week": "Day",
        "start_time": "Start Time",
        "duration_hours": "Duration Hours",
        "contract_hours": "Contract Hours",
    }
    index = _validate_columns(header, mapped)

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[index["student_name"]]:
            continue

        therapist_name = str(row[index["therapist_name"]]).strip()
        therapist = Therapist.query.filter_by(name=therapist_name).first() or Therapist(name=therapist_name)
        db.session.add(therapist)
        db.session.flush()

        student_name = str(row[index["student_name"]]).strip()
        student = Student.query.filter_by(name=student_name).first()
        if not student:
            student = Student(
                name=student_name,
                assigned_therapist_id=therapist.id,
                contract_hours_per_week=float(row[index["contract_hours"]] or 1),
            )
            db.session.add(student)
            db.session.flush()

        day = DAY_MAP[str(row[index["day_of_week"]]).strip().lower()]
        st_raw = row[index["start_time"]]
        start_time = st_raw if hasattr(st_raw, "hour") else datetime.strptime(str(st_raw), "%H:%M").time()
        duration = float(row[index["duration_hours"]])
        schedule = RegularSchedule(
            student_id=student.id,
            therapist_id=therapist.id,
            day_of_week=day,
            start_time=start_time,
            end_time=_end_time_from_duration(start_time, duration),
            duration_hours=duration,
        )
        db.session.add(schedule)
        inserted += 1
    db.session.commit()
    return inserted


def import_admin_staff(path: str, column_map: dict[str, str] | None = None) -> int:
    wb = load_workbook(path)
    ws = wb.active
    header = _sheet_header(ws)
    mapped = column_map or {"name": "Admin Name", "active": "Active"}
    index = _validate_columns(header, mapped)

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[index["name"]]:
            continue
        name = str(row[index["name"]]).strip()
        existing = AdminStaff.query.filter_by(name=name).first()
        if existing:
            continue
        active_raw = str(row[index["active"]]).strip().lower()
        active = active_raw in {"1", "true", "yes", "y", "active"}
        db.session.add(AdminStaff(name=name, active=active))
        inserted += 1
    db.session.commit()
    return inserted


def import_therapists(path: str, column_map: dict[str, str] | None = None) -> int:
    wb = load_workbook(path)
    ws = wb.active
    header = _sheet_header(ws)
    mapped = column_map or {"name": "Therapist Name", "active": "Active"}
    index = _validate_columns(header, mapped)

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[index["name"]]:
            continue
        name = str(row[index["name"]]).strip()
        existing = Therapist.query.filter_by(name=name).first()
        if existing:
            continue
        active_raw = str(row[index["active"]]).strip().lower()
        active = active_raw in {"1", "true", "yes", "y", "active"}
        db.session.add(Therapist(name=name, active=active))
        inserted += 1
    db.session.commit()
    return inserted


def _save_workbook(wb: Workbook, path: str) -> str:
    target = Path(path).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(target))
    return str(target)


def export_attendance_summary(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Date", "Student", "Therapist", "Start", "Duration", "Status", "Source"])
    for s in AttendanceSession.query.order_by(AttendanceSession.session_date).all():
        ws.append([
            s.session_date.isoformat(),
            s.student.name,
            s.therapist.name,
            s.start_time.strftime("%H:%M"),
            s.duration_hours,
            s.status,
            s.source_type,
        ])
    return _save_workbook(wb, path)


def export_admin_attendance(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Attendance"
    ws.append(["Date", "Admin", "Status", "Shift", "Hours"])
    for a in AdminAttendance.query.order_by(AdminAttendance.attendance_date).all():
        ws.append([a.attendance_date.isoformat(), a.admin.name, a.status, a.shift_label, a.hours_worked])
    return _save_workbook(wb, path)


def export_therapist_weekly_hours(path: str, start_date: date, end_date: date) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Therapist Weekly Hours"
    ws.append(["Start", "End", "Therapist", "Hours"])
    totals = weekly_therapist_hours(start_date, end_date)
    for therapist_id, hours in totals.items():
        therapist = Therapist.query.get(therapist_id)
        ws.append([start_date.isoformat(), end_date.isoformat(), therapist.name if therapist else therapist_id, hours])
    return _save_workbook(wb, path)


def export_billing_advices(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Advice"
    ws.append([
        "Student",
        "Cycle Start",
        "Cycle End",
        "Issue",
        "Due",
        "Regular Rendered Hours",
        "Make-up Rendered Hours",
        "Total Rendered Hours",
        "Billable Hours",
        "Non-billable Hours",
        "Subtotal",
        "Old Balance",
        "Required Deposit",
        "Assessment Deposit",
        "Total Due",
        "Status",
    ])
    replaced_ids = {
        row[0]
        for row in db.session.query(SessionOverride.original_session_id)
        .filter(SessionOverride.original_session_id.isnot(None))
        .all()
    }
    for advice in BillingAdvice.query.order_by(BillingAdvice.id).all():
        cycle = advice.billing_cycle
        sessions = AttendanceSession.query.filter(
            AttendanceSession.student_id == advice.student_id,
            AttendanceSession.session_date >= cycle.start_date,
            AttendanceSession.session_date <= cycle.end_date,
            AttendanceSession.status.in_(["Present", "Make-up", "Rescheduled", "Non-billable"]),
        ).all()

        regular_hours = 0.0
        makeup_hours = 0.0
        billable_hours = 0.0
        non_billable_hours = 0.0
        for sess in sessions:
            if sess.id in replaced_ids:
                continue
            if sess.status in {"Present", "Make-up", "Rescheduled"}:
                billable_hours += sess.duration_hours
                if sess.session_type == "makeup":
                    makeup_hours += sess.duration_hours
                else:
                    regular_hours += sess.duration_hours
            elif sess.status == "Non-billable":
                non_billable_hours += sess.duration_hours

        total_rendered_hours = regular_hours + makeup_hours + non_billable_hours
        ws.append([
            advice.student.name,
            cycle.start_date.isoformat(),
            cycle.end_date.isoformat(),
            cycle.issue_date.isoformat(),
            cycle.due_date.isoformat(),
            round(regular_hours, 2),
            round(makeup_hours, 2),
            round(total_rendered_hours, 2),
            round(billable_hours, 2),
            round(non_billable_hours, 2),
            advice.subtotal_sessions,
            advice.old_balance,
            advice.required_deposit_charge,
            advice.assessment_deposit_charge,
            advice.total_due,
            advice.status,
        ])
    return _save_workbook(wb, path)


def export_payment_ledger(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(["Payment ID", "Date", "Student", "Amount", "Allocation Type", "Allocation Amount", "Advice ID", "Notes"])
    for p in Payment.query.order_by(Payment.payment_date, Payment.id).all():
        if p.allocations:
            for alloc in p.allocations:
                ws.append([
                    p.id,
                    p.payment_date.isoformat(),
                    p.student.name,
                    p.amount,
                    alloc.allocation_type,
                    alloc.amount,
                    alloc.billing_advice_id,
                    p.notes,
                ])
        else:
            ws.append([p.id, p.payment_date.isoformat(), p.student.name, p.amount, "", 0, "", p.notes])
    return _save_workbook(wb, path)


def _export_deposit_payment_history(path: str, allocation_type: str, student_id: int | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deposit Payments"
    ws.append([
        "Payment ID",
        "Payment Date",
        "Client/Guardian",
        "Student",
        "Amount Paid",
        "Received By",
        "Purpose",
        "Billing Period Start",
        "Billing Period End",
        "Mode of Transfer",
        "Notes",
        "Allocation Type",
        "Allocation Amount",
        "Billing Advice ID",
        "Billing Cycle ID",
    ])

    query = PaymentAllocation.query.filter_by(allocation_type=allocation_type).join(Payment, PaymentAllocation.payment_id == Payment.id)
    if student_id:
        query = query.filter(Payment.student_id == student_id)

    for alloc in query.order_by(Payment.payment_date, Payment.id).all():
        payment = alloc.payment
        advice = BillingAdvice.query.get(alloc.billing_advice_id) if alloc.billing_advice_id else None
        ws.append([
            payment.id,
            payment.payment_date.isoformat(),
            payment.client_guardian_name,
            payment.student.name,
            payment.amount,
            payment.received_by_admin.name if payment.received_by_admin else "",
            payment.purpose,
            payment.billing_period_start.isoformat() if payment.billing_period_start else "",
            payment.billing_period_end.isoformat() if payment.billing_period_end else "",
            payment.mode_of_transfer,
            payment.notes,
            alloc.allocation_type,
            alloc.amount,
            alloc.billing_advice_id or "",
            advice.billing_cycle_id if advice else "",
        ])
    return _save_workbook(wb, path)


def export_required_deposit_payment_history(path: str, student_id: int | None = None) -> str:
    return _export_deposit_payment_history(path, allocation_type="required_deposit", student_id=student_id)


def export_assessment_deposit_payment_history(path: str, student_id: int | None = None) -> str:
    return _export_deposit_payment_history(path, allocation_type="assessment_deposit", student_id=student_id)

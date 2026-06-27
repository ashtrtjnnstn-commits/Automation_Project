from __future__ import annotations

from datetime import date, datetime, time, timedelta

import pytest
from openpyxl import Workbook, load_workbook

from app import create_app
from app.models import (
    AdminStaff,
    AssessmentDepositLedger,
    AttendanceSession,
    AuditLog,
    BillingAdvice,
    MonthlyPaymentArchive,
    PaymentAllocation,
    Payment,
    RedBillingNotice,
    RegularSchedule,
    RequiredDepositLedger,
    SessionOverride,
    Student,
    Supervisor,
    Therapist,
    WeeklyReportArchive,
    db,
)
from app.services.attendance_service import create_makeup_session, generate_monthly_sessions, missed_recovery_summary, weekly_student_hours, weekly_therapist_hours
from app.services.billing_service import WEEKDAY_RATE, billing_hours_breakdown, generate_billing_advices_for_cycle, generate_billing_cycles_for_range
from app.services.import_export_service import import_students_and_schedules
from app.services.import_export_service import (
    export_assessment_deposit_payment_history,
    export_billing_advices,
    export_required_deposit_payment_history,
)
from app.services.payment_service import record_payment
from app.utils.backup_utils import backup_sqlite_database, restore_sqlite_backup


def setup_basic():
    t = Therapist(name="T")
    s = Student(name="S", contract_hours_per_week=2)
    db.session.add_all([t, s])
    db.session.flush()
    s.assigned_therapist_id = t.id
    db.session.add(RegularSchedule(student_id=s.id, therapist_id=t.id, day_of_week=0, start_time=time(9, 0), duration_hours=1))
    db.session.commit()
    return s, t


def mark_rendered(student_id: int | None = None):
    q = AttendanceSession.query
    if student_id:
        q = q.filter_by(student_id=student_id)
    for sess in q.all():
        sess.status = "Present"
    db.session.commit()


def test_monthly_session_generation(session):
    setup_basic()
    count = generate_monthly_sessions(2026, 1)
    assert count > 0


def test_makeup_session_creation(session):
    s, t = setup_basic()
    session_obj = create_makeup_session(s.id, t.id, date(2026, 1, 2), time(10, 0), 1.0)
    assert session_obj.session_type == "makeup"
    assert session_obj.status == ""


def test_weekly_hours_calculation(session):
    s, t = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    student_hours = weekly_student_hours(date(2026, 1, 1), date(2026, 1, 7))
    therapist_hours = weekly_therapist_hours(date(2026, 1, 1), date(2026, 1, 7))
    assert student_hours[s.id] == therapist_hours[t.id]


def test_billing_cycle_generation_and_due_date(session):
    cycles = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 31))
    assert len(cycles) == 3
    assert (cycles[0].due_date - cycles[0].issue_date).days == 5


def test_weekday_weekend_rate_billing(session):
    t = Therapist(name="T2")
    s = Student(name="S2", contract_hours_per_week=2, assigned_therapist=t)
    db.session.add_all([t, s])
    db.session.flush()
    db.session.add(RegularSchedule(student_id=s.id, therapist_id=t.id, day_of_week=4, start_time=time(9, 0), duration_hours=1))
    db.session.add(RegularSchedule(student_id=s.id, therapist_id=t.id, day_of_week=5, start_time=time(9, 0), duration_hours=1))
    db.session.commit()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advices = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)
    advice = advices[0]
    assert advice.subtotal_sessions >= WEEKDAY_RATE


def test_required_deposit_stops_after_four_cycles(session):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycles = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 3, 31))
    for c in cycles[:5]:
        generate_billing_advices_for_cycle(c.id, student_id=s.id)
    advices = BillingAdvice.query.filter_by(student_id=s.id).order_by(BillingAdvice.id).all()
    non_zero = [a for a in advices if a.required_deposit_charge > 0]
    assert len(non_zero) <= 4


def test_assessment_deposit_tracking(session):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.assessment_deposit_charge == 2500.0


def test_assessment_deposit_defaults_enabled_for_existing_compatibility(session):
    s = Student(name="Assessment Policy Default", contract_hours_per_week=2)
    db.session.add(s)
    db.session.commit()
    assert s.assessment_deposit_enabled is True


def test_assessment_deposit_disabled_prevents_billing_charge(session):
    s, _ = setup_basic()
    s.assessment_deposit_enabled = False
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.assessment_deposit_charge == 0.0


def test_required_deposit_charge_respects_paid_amount(session):
    s, _ = setup_basic()
    s.required_deposit_total = 10000
    s.required_deposit_billed = 2500
    s.required_deposit_paid = 9500
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.required_deposit_charge == 500


def test_assessment_deposit_charge_respects_paid_amount(session):
    s, _ = setup_basic()
    s.assessment_deposit_total = 5000
    s.assessment_deposit_billed = 0
    s.assessment_deposit_paid = 4900
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.assessment_deposit_charge == 100


def test_billing_normalizes_negative_student_finance_values(session):
    s, _ = setup_basic()
    s.required_deposit_total = -100
    s.required_deposit_billed = -25
    s.required_deposit_paid = -10
    s.assessment_deposit_total = -100
    s.assessment_deposit_billed = -20
    s.assessment_deposit_paid = -5
    s.overpayment_credit = -7
    db.session.commit()

    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    db.session.refresh(s)
    assert s.required_deposit_total >= 0
    assert s.required_deposit_billed >= 0
    assert s.required_deposit_paid >= 0
    assert s.assessment_deposit_total >= 0
    assert s.assessment_deposit_billed >= 0
    assert s.assessment_deposit_paid >= 0
    assert s.overpayment_credit >= 0
    assert advice.required_deposit_charge >= 0
    assert advice.assessment_deposit_charge >= 0


def test_deposit_charge_helpers_handle_none_finance_values_without_crash(session):
    s, _ = setup_basic()
    s.required_deposit_total = None
    s.required_deposit_billed = None
    s.required_deposit_paid = None
    s.assessment_deposit_total = None
    s.assessment_deposit_billed = None
    s.assessment_deposit_paid = None
    s.overpayment_credit = None
    from app.services.billing_service import _assessment_deposit_charge, _required_deposit_charge

    with db.session.no_autoflush:
        req = _required_deposit_charge(s)
        ass = _assessment_deposit_charge(s)
    assert req >= 0
    assert ass >= 0


def test_billing_excludes_replaced_original_session(session):
    s, t = setup_basic()
    original = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        source_type="generated",
    )
    replacement = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 6),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Rescheduled",
        source_type="manual",
    )
    db.session.add_all([original, replacement])
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=original.id, new_session_id=replacement.id, override_type="reschedule"))
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.subtotal_sessions == 550.0


def test_partial_payment_allocation(session):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    before = advice.total_due
    record_payment(s.id, 1000, date(2026, 1, 16))
    refreshed = BillingAdvice.query.get(advice.id)
    assert refreshed.total_due == max(before - 1000, 0)


def test_overpayment_carry_forward(session):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    record_payment(s.id, advice.total_due + 500, date(2026, 1, 16))
    db.session.refresh(s)
    assert s.overpayment_credit == 500


def test_import_validates_required_columns(session, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Student Name", "Therapist"])
    ws.append(["Example", "Therapist A"])
    path = tmp_path / "bad_import.xlsx"
    wb.save(path)

    with pytest.raises(ValueError):
        import_students_and_schedules(str(path))


def test_daily_schedule_updates_attendance_record(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    target = AttendanceSession.query.filter_by(student_id=s.id).first()

    response = client.post(
        "/attendance/daily",
        data={
            "selected_date": target.session_date.isoformat(),
            "session_ids": [str(target.id)],
            f"status_{target.id}": "Present",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    db.session.refresh(target)
    assert target.status == "Present"


def test_monthly_attendance_reflects_daily_status(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    target = AttendanceSession.query.filter_by(student_id=s.id).first()
    target.status = "Absent"
    db.session.commit()

    response = client.get(f"/attendance?year=2026&month=1&student_id={s.id}")
    assert response.status_code == 200
    assert b"Absent" in response.data


def test_payment_dashboard_record_creation(session, client):
    s, _ = setup_basic()
    admin = AdminStaff(name="Receiver A")
    db.session.add(admin)
    db.session.commit()

    response = client.post(
        "/payments",
        data={
            "payment_date": "2026-01-10",
            "client_guardian_name": "Juan Dela Cruz",
            "student_id": str(s.id),
            "purpose": "Therapy",
            "billing_period_start": "2026-01-01",
            "billing_period_end": "2026-01-15",
            "total_hours_rendered": "2",
            "amount": "1000",
            "received_by_admin_id": str(admin.id),
            "overpayment_amount": "12.5",
            "balance_after_payment": "88.0",
            "mode_of_transfer": "GCash",
            "notes": "test",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    payment = Payment.query.order_by(Payment.id.desc()).first()
    assert payment.client_guardian_name == "Juan Dela Cruz"
    assert payment.mode_of_transfer == "GCash"
    assert payment.overpayment_amount == 12.5
    assert payment.balance_after_payment == 88.0


def test_monthly_payment_archive_retrieval(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, client_guardian_name="A"))
    db.session.commit()

    res_archive = client.post("/payments/tracker", data={"action": "archive", "archive_month": "1", "archive_year": "2026"}, follow_redirects=True)
    assert res_archive.status_code == 200

    res_view = client.get("/payments/tracker?view=archived&month=1&year=2026")
    assert res_view.status_code == 200
    assert b"2026-01-10" in res_view.data


def test_billing_page_requires_student_selection(session, client):
    response = client.post(
        "/billing",
        data={"start_date": "2026-01-01", "end_date": "2026-01-15", "student_id": ""},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b"Please select a student first." in response.data


def test_payments_tracker_filters_current_month(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, client_guardian_name="A", is_archived=False))
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 2, 10), amount=700, client_guardian_name="B", is_archived=False))
    db.session.commit()

    res = client.get("/payments/tracker?view=active&month=1&year=2026")
    assert res.status_code == 200
    assert b"2026-01-10" in res.data
    assert b"2026-02-10" not in res.data


def test_edit_payment_entry(session, client):
    s, _ = setup_basic()
    admin = AdminStaff(name="Receiver Edit")
    db.session.add(admin)
    db.session.flush()
    p = Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, client_guardian_name="Before", mode_of_transfer="Cash")
    db.session.add(p)
    db.session.commit()

    res = client.post(
        f"/payments/{p.id}/edit",
        data={
            "payment_date": "2026-01-12",
            "client_guardian_name": "After",
            "student_id": str(s.id),
            "purpose": "Partial Payment",
            "billing_period_start": "2026-01-01",
            "billing_period_end": "2026-01-15",
            "total_hours_rendered": "3",
            "amount": "750",
            "received_by_admin_id": str(admin.id),
            "overpayment_amount": "10",
            "balance_after_payment": "20",
            "mode_of_transfer": "GCash",
            "notes": "edited",
        },
        follow_redirects=True,
    )
    assert res.status_code == 200
    db.session.refresh(p)
    assert p.payment_date == date(2026, 1, 12)
    assert p.client_guardian_name == "After"
    assert p.purpose == "Partial Payment"
    assert p.amount == 750
    assert p.mode_of_transfer == "GCash"


def test_master_data_student_crud_page(session, client):
    res_create = client.post(
        "/master-data/students",
        data={"action": "create", "name": "New Student", "contract_hours_per_week": "4", "overpayment_credit": "0", "active": "1"},
        follow_redirects=True,
    )
    assert res_create.status_code == 200
    student = Student.query.filter_by(name="New Student").first()
    assert student is not None

    res_edit = client.post(
        "/master-data/students",
        data={"action": "edit", "student_id": str(student.id), "name": "Updated Student", "contract_hours_per_week": "5", "overpayment_credit": "1.5", "active": "0"},
        follow_redirects=True,
    )
    assert res_edit.status_code == 200
    db.session.refresh(student)
    assert student.name == "Updated Student"
    assert student.active is False


def test_master_data_schedule_validation_end_time(session, client):
    s, t = setup_basic()
    response = client.post(
        "/master-data/schedules",
        data={
            "action": "create",
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "09:00",
            "active": "1",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b"End time must be after start time." in response.data


def _setup_generated_schedule_sync_case():
    s, t = setup_basic()
    sched = RegularSchedule.query.filter_by(student_id=s.id).first()
    generate_monthly_sessions(2026, 3)
    return s, t, sched


def test_schedule_sync_keeps_past_generated_attendance_unchanged(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    past = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 16), source_type="generated").first()
    assert past is not None

    client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
            "sync_confirmed": "1",
        },
        follow_redirects=True,
    )

    unchanged = AttendanceSession.query.get(past.id)
    assert unchanged is not None
    assert unchanged.start_time == time(9, 0)


def test_schedule_sync_updates_future_blank_generated_from_effective_date(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()

    client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
            "sync_confirmed": "1",
        },
        follow_redirects=True,
    )

    removed_old = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 30), start_time=time(9, 0), source_type="generated").first()
    added_new = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 24), start_time=time(10, 0), source_type="generated").first()
    assert removed_old is None
    assert added_new is not None


def test_schedule_sync_preserves_recorded_future_attendance(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    recorded = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 23), start_time=time(9, 0), source_type="generated").first()
    recorded.status = "Present"
    db.session.commit()

    client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
            "sync_confirmed": "1",
        },
        follow_redirects=True,
    )

    still_recorded = AttendanceSession.query.get(recorded.id)
    assert still_recorded is not None
    assert still_recorded.status == "Present"


def test_schedule_sync_preserves_manual_overrides(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    manual = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 3, 24),
        start_time=time(10, 0),
        end_time=time(11, 0),
        duration_hours=1,
        session_type="makeup",
        source_type="manual",
        status="",
    )
    db.session.add(manual)
    db.session.commit()

    client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
            "sync_confirmed": "1",
        },
        follow_redirects=True,
    )

    kept_manual = AttendanceSession.query.get(manual.id)
    generated_same_slot = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 24), start_time=time(10, 0), source_type="generated").first()
    assert kept_manual is not None
    assert kept_manual.source_type == "manual"
    assert generated_same_slot is None


def test_schedule_sync_removes_old_future_blank_dates_and_adds_new_safe_dates(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    old_blank = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 30), start_time=time(9, 0), source_type="generated").first()
    assert old_blank is not None

    res = client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
            "sync_confirmed": "1",
        },
        follow_redirects=True,
    )

    assert b"Apply this permanent schedule change to future unmarked generated attendance from the effective date onward." in res.data
    assert AttendanceSession.query.get(old_blank.id) is None
    assert AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 24), start_time=time(10, 0), source_type="generated").first() is not None


def test_schedule_sync_preview_step_appears_with_counts(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    recorded = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 23), source_type="generated").first()
    recorded.status = "Present"
    original_for_override = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 30), source_type="generated").first()
    manual = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 3, 30),
        start_time=time(11, 0),
        end_time=time(12, 0),
        duration_hours=1,
        session_type="makeup",
        source_type="manual",
        status="",
    )
    db.session.add(manual)
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=original_for_override.id, new_session_id=manual.id, override_type="makeup"))
    db.session.commit()

    preview = client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
        },
        follow_redirects=True,
    )

    assert b"Schedule Sync Preview" in preview.data
    assert b"future blank generated session(s) will be removed" in preview.data
    assert b"new future generated session(s) will be added" in preview.data
    assert b"recorded session(s) will be preserved" in preview.data
    assert b"override-related session(s) will be preserved" in preview.data


def test_cancel_schedule_sync_preview_does_not_apply_changes(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    old_blank = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 30), start_time=time(9, 0), source_type="generated").first()

    client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
        },
        follow_redirects=True,
    )

    cancel = client.post("/master-data/schedules", data={"sync_cancel": "1"}, follow_redirects=True)
    assert b"No attendance changes were applied." in cancel.data
    assert AttendanceSession.query.get(old_blank.id) is not None
    assert AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 24), start_time=time(10, 0), source_type="generated").first() is None


def test_confirm_schedule_sync_preview_applies_and_shows_human_feedback(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    old_blank = AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 30), start_time=time(9, 0), source_type="generated").first()

    client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
        },
        follow_redirects=True,
    )

    confirmed = client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "sync_confirmed": "1",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-03-18",
            "apply_future_sync": "1",
        },
        follow_redirects=True,
    )

    assert b"Schedule updated successfully." in confirmed.data
    assert b"outdated future sessions removed." in confirmed.data
    assert b"new future sessions added." in confirmed.data
    assert b"Recorded attendance and overrides were preserved." in confirmed.data
    assert AttendanceSession.query.get(old_blank.id) is None
    assert AttendanceSession.query.filter_by(student_id=s.id, session_date=date(2026, 3, 24), start_time=time(10, 0), source_type="generated").first() is not None


def test_schedule_sync_ui_helper_texts_and_default_on_render(session, client):
    setup_basic()
    page = client.get("/master-data/schedules")
    assert b"Changes will apply only to future unmarked sessions starting this date." in page.data
    assert b"This will not change past or recorded attendance." in page.data
    assert b'<option value=\"1\" selected>Yes</option>' in page.data


def test_schedule_sync_preview_shows_past_date_warning(session, client):
    s, t, sched = _setup_generated_schedule_sync_case()
    preview = client.post(
        "/master-data/schedules",
        data={
            "action": "edit",
            "schedule_id": str(sched.id),
            "student_id": str(s.id),
            "therapist_id": str(t.id),
            "day_of_week": "1",
            "start_time": "10:00",
            "end_time": "11:00",
            "active": "1",
            "effective_from": "2026-01-01",
            "apply_future_sync": "1",
        },
        follow_redirects=True,
    )
    assert b"Past dates will not be modified. Only future unmarked sessions will be affected." in preview.data


def test_weekly_report_archive_creation_and_view(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)

    res = client.post("/reports/weekly", data={"date": "2026-01-05", "action": "archive_week", "note": "snapshot"}, follow_redirects=True)
    assert res.status_code == 200
    assert b"Weekly report for 2026-01-05 to 2026-01-11 archived successfully." in res.data

    # locate archive link and open list page
    list_res = client.get("/reports/weekly?date=2026-01-05")
    assert list_res.status_code == 200
    assert b"Archived Weekly Reports" in list_res.data


def test_weekly_report_archive_prevents_duplicate(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)

    client.post("/reports/weekly", data={"date": "2026-01-05", "action": "archive_week", "note": "a"}, follow_redirects=True)
    res2 = client.post("/reports/weekly", data={"date": "2026-01-05", "action": "archive_week", "note": "b"}, follow_redirects=True)
    assert res2.status_code == 200
    assert b"already archived" in res2.data


def test_required_deposit_export_includes_expected_row_and_columns_for_student(session, tmp_path):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    record_payment(s.id, advice.required_deposit_charge, date(2026, 1, 16), client_guardian_name="Guardian A", purpose="Required Deposit", mode_of_transfer="GCash")

    output = export_required_deposit_payment_history(str(tmp_path / "required_one.xlsx"), student_id=s.id)
    wb = load_workbook(output)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Payment Date" in headers
    assert "Client/Guardian" in headers
    assert "Allocation Amount" in headers
    assert ws.max_row >= 2
    assert ws.cell(row=2, column=headers.index("Student") + 1).value == s.name


def test_required_deposit_export_supports_all_students(session, tmp_path):
    s1, _ = setup_basic()
    s2 = Student(name="S-2", contract_hours_per_week=2)
    db.session.add(s2)
    db.session.commit()

    c = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    a1 = generate_billing_advices_for_cycle(c.id, student_id=s1.id)[0]
    a2 = generate_billing_advices_for_cycle(c.id, student_id=s2.id)[0]
    record_payment(s1.id, a1.required_deposit_charge, date(2026, 1, 16), purpose="Required Deposit")
    record_payment(s2.id, a2.required_deposit_charge, date(2026, 1, 16), purpose="Required Deposit")

    output = export_required_deposit_payment_history(str(tmp_path / "required_all.xlsx"))
    ws = load_workbook(output).active
    names = {ws.cell(row=i, column=4).value for i in range(2, ws.max_row + 1)}
    assert s1.name in names
    assert s2.name in names


def test_assessment_deposit_export_includes_expected_row_and_columns_for_student(session, tmp_path):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    record_payment(s.id, advice.assessment_deposit_charge, date(2026, 1, 16), client_guardian_name="Guardian B", purpose="Assessment Deposit", mode_of_transfer="Bank Transfer")

    output = export_assessment_deposit_payment_history(str(tmp_path / "assessment_one.xlsx"), student_id=s.id)
    wb = load_workbook(output)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Payment Date" in headers
    assert "Mode of Transfer" in headers
    assert "Allocation Amount" in headers
    assert ws.max_row >= 2
    assert ws.cell(row=2, column=headers.index("Student") + 1).value == s.name


def test_assessment_deposit_export_supports_all_students(session, tmp_path):
    s1, _ = setup_basic()
    s2 = Student(name="S-3", contract_hours_per_week=2)
    db.session.add(s2)
    db.session.commit()

    c = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    a1 = generate_billing_advices_for_cycle(c.id, student_id=s1.id)[0]
    a2 = generate_billing_advices_for_cycle(c.id, student_id=s2.id)[0]
    record_payment(s1.id, a1.assessment_deposit_charge, date(2026, 1, 16), purpose="Assessment Deposit")
    record_payment(s2.id, a2.assessment_deposit_charge, date(2026, 1, 16), purpose="Assessment Deposit")

    output = export_assessment_deposit_payment_history(str(tmp_path / "assessment_all.xlsx"))
    ws = load_workbook(output).active
    names = {ws.cell(row=i, column=4).value for i in range(2, ws.max_row + 1)}
    assert s1.name in names
    assert s2.name in names


def test_billing_export_includes_regular_and_makeup_hours_columns(session, tmp_path):
    s, t = setup_basic()
    original = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Cancelled",
        session_type="regular",
        source_type="generated",
    )
    makeup = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 6),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="makeup",
        source_type="manual",
    )
    db.session.add_all([original, makeup])
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=original.id, new_session_id=makeup.id, override_type="makeup"))
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    generate_billing_advices_for_cycle(cycle.id, student_id=s.id)

    output = export_billing_advices(str(tmp_path / "billing.xlsx"))
    ws = load_workbook(output).active
    headers = [c.value for c in ws[1]]
    assert "Regular Rendered Hours" in headers
    assert "Make-up Rendered Hours" in headers
    reg_col = headers.index("Regular Rendered Hours") + 1
    make_col = headers.index("Make-up Rendered Hours") + 1
    assert ws.cell(row=2, column=reg_col).value == 0
    assert ws.cell(row=2, column=make_col).value == 1


def test_missed_hours_summary_computation(session):
    s, t = setup_basic()
    missed = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Absent",
        session_type="regular",
        source_type="generated",
    )
    recovered = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 6),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=0.5,
        status="Present",
        session_type="makeup",
        source_type="manual",
    )
    db.session.add_all([missed, recovered])
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=missed.id, new_session_id=recovered.id, override_type="makeup"))
    db.session.commit()

    summary = missed_recovery_summary(student_id=s.id)
    assert summary["missed_hours"] == 1.0
    assert summary["recovered_makeup_hours"] == 0.5
    assert summary["remaining_missed_hours"] == 0.5


def test_missed_recovery_is_student_specific(session):
    t = Therapist(name="T-main")
    s1 = Student(name="S-main", contract_hours_per_week=2, assigned_therapist=t)
    s2 = Student(name="S-other", contract_hours_per_week=2, assigned_therapist=t)
    db.session.add_all([t, s1, s2])
    db.session.flush()

    s1_missed = AttendanceSession(
        student_id=s1.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Absent",
        session_type="regular",
        source_type="generated",
    )
    s2_makeup = AttendanceSession(
        student_id=s2.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 6),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="makeup",
        source_type="manual",
    )
    db.session.add_all([s1_missed, s2_makeup])
    db.session.flush()
    # Link intentionally mismatched to ensure recovery still remains student-specific.
    db.session.add(SessionOverride(original_session_id=s1_missed.id, new_session_id=s2_makeup.id, override_type="makeup"))
    db.session.commit()

    s1_summary = missed_recovery_summary(student_id=s1.id)
    assert s1_summary["missed_hours"] == 1.0
    assert s1_summary["recovered_makeup_hours"] == 0.0
    assert s1_summary["remaining_missed_hours"] == 1.0


def test_student_b_makeup_does_not_reduce_student_a_remaining(session):
    t = Therapist(name="T-iso")
    s1 = Student(name="S-A", contract_hours_per_week=2, assigned_therapist=t)
    s2 = Student(name="S-B", contract_hours_per_week=2, assigned_therapist=t)
    db.session.add_all([t, s1, s2])
    db.session.flush()

    s1_missed = AttendanceSession(
        student_id=s1.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Cancelled",
        session_type="regular",
        source_type="generated",
    )
    s2_missed = AttendanceSession(
        student_id=s2.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(10, 0),
        end_time=time(11, 0),
        duration_hours=1,
        status="Absent",
        session_type="regular",
        source_type="generated",
    )
    s2_makeup = AttendanceSession(
        student_id=s2.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 6),
        start_time=time(10, 0),
        end_time=time(11, 0),
        duration_hours=1,
        status="Present",
        session_type="makeup",
        source_type="manual",
    )
    db.session.add_all([s1_missed, s2_missed, s2_makeup])
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=s2_missed.id, new_session_id=s2_makeup.id, override_type="makeup"))
    db.session.commit()

    s1_summary = missed_recovery_summary(student_id=s1.id)
    s2_summary = missed_recovery_summary(student_id=s2.id)
    assert s1_summary["remaining_missed_hours"] == 1.0
    assert s2_summary["remaining_missed_hours"] == 0.0




def test_dashboard_shows_cumulative_outstanding_makeup_obligations(session, client):
    s, t = setup_basic()
    missed = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Cancelled",
        session_type="regular",
        source_type="generated",
    )
    recovered = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 2, 6),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=0.5,
        status="Present",
        session_type="makeup",
        source_type="manual",
    )
    db.session.add_all([missed, recovered])
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=missed.id, new_session_id=recovered.id, override_type="makeup"))
    db.session.commit()

    res = client.get("/")
    assert res.status_code == 200
    assert b"Outstanding Make-up Obligations" in res.data
    assert b"Total Missed Hours To Date" in res.data
    assert b"Total Recovered Make-up Hours To Date" in res.data
    assert b"Outstanding Make-up Hours" in res.data
    assert b"1.0" in res.data
    assert b"0.5" in res.data


def test_weekly_reports_no_longer_shows_missed_vs_makeup_block(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)

    res = client.get("/reports/weekly?date=2026-01-05")
    assert res.status_code == 200
    assert b"Missed vs Make-up Recovery" not in res.data

def test_rendered_billable_sessions_produce_non_zero_subtotal(session):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    sess = AttendanceSession.query.filter(
        AttendanceSession.student_id == s.id,
        AttendanceSession.session_date >= cycle.start_date,
        AttendanceSession.session_date <= cycle.end_date,
    ).first()
    if not sess:
        generate_monthly_sessions(2026, 1)
        sess = AttendanceSession.query.filter(
            AttendanceSession.student_id == s.id,
            AttendanceSession.session_date >= cycle.start_date,
            AttendanceSession.session_date <= cycle.end_date,
        ).first()
    sess.status = "Present"
    db.session.commit()

    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.subtotal_sessions > 0


def test_billing_hours_breakdown_matches_effective_sessions(session):
    s, t = setup_basic()
    original = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 5),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Rescheduled",
        session_type="regular",
        source_type="generated",
    )
    makeup = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 6),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="makeup",
        source_type="manual",
    )
    non_billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 7),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=0.5,
        status="Non-billable",
        session_type="regular",
        source_type="manual",
    )
    db.session.add_all([original, makeup, non_billable])
    db.session.flush()
    db.session.add(SessionOverride(original_session_id=original.id, new_session_id=makeup.id, override_type="makeup"))
    db.session.commit()

    hours = billing_hours_breakdown(s.id, date(2026, 1, 1), date(2026, 1, 15))
    assert hours["regular_rendered_hours"] == 0.0
    assert hours["makeup_rendered_hours"] == 1.0
    assert hours["billable_hours"] == 1.0
    assert hours["non_billable_hours"] == 0.5
    assert hours["total_rendered_hours"] == 1.5


def test_non_billable_in_rendered_not_billed_totals(session):
    s, t = setup_basic()
    billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 8),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="regular",
        source_type="generated",
    )
    non_billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 9),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Non-billable",
        session_type="regular",
        source_type="manual",
    )
    db.session.add_all([billable, non_billable])
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    hours = billing_hours_breakdown(s.id, cycle.start_date, cycle.end_date)

    assert hours["total_rendered_hours"] == 2.0
    assert hours["billable_hours"] == 1.0
    assert advice.subtotal_sessions == 550.0


def test_exact_required_deposit_payment_has_zero_overpayment(session):
    s, _ = setup_basic()
    s.required_deposit_total = 1000
    s.required_deposit_paid = 0
    db.session.commit()

    payment = record_payment(s.id, 1000, date(2026, 1, 10), purpose="Required Deposit")
    db.session.refresh(s)
    assert payment.overpayment_amount == 0
    assert s.overpayment_credit == 0
    assert payment.balance_after_payment == 0
    assert s.required_deposit_paid == 1000


def test_exact_assessment_deposit_payment_has_zero_overpayment(session):
    s, _ = setup_basic()
    s.assessment_deposit_total = 500
    s.assessment_deposit_paid = 0
    db.session.commit()

    payment = record_payment(s.id, 500, date(2026, 1, 10), purpose="Assessment Deposit")
    db.session.refresh(s)
    assert payment.overpayment_amount == 0
    assert s.overpayment_credit == 0
    assert payment.balance_after_payment == 0
    assert s.assessment_deposit_paid == 500


def test_deposit_true_excess_only_excess_becomes_overpayment(session):
    s, _ = setup_basic()
    s.required_deposit_total = 1000
    s.required_deposit_paid = 0
    db.session.commit()

    payment = record_payment(s.id, 1200, date(2026, 1, 10), purpose="Required Deposit")
    db.session.refresh(s)
    assert payment.overpayment_amount == 200
    assert s.overpayment_credit == 200
    assert payment.balance_after_payment == 0


def test_partial_deposit_payment_updates_remaining_balance(session):
    s, _ = setup_basic()
    s.assessment_deposit_total = 1000
    s.assessment_deposit_paid = 0
    db.session.commit()

    payment = record_payment(s.id, 300, date(2026, 1, 10), purpose="Assessment Deposit")
    db.session.refresh(s)
    assert payment.overpayment_amount == 0
    assert payment.balance_after_payment == 700
    assert s.assessment_deposit_paid == 300


def test_billing_advice_does_not_recharge_fully_paid_deposits(session):
    s, t = setup_basic()
    s.required_deposit_total = 1000
    s.assessment_deposit_total = 500
    s.required_deposit_paid = 1000
    s.assessment_deposit_paid = 500
    s.required_deposit_billed = 0
    s.assessment_deposit_billed = 0
    db.session.add(RequiredDepositLedger(student_id=s.id, entry_type="paid", amount=1000))
    db.session.add(AssessmentDepositLedger(student_id=s.id, entry_type="paid", amount=500))

    billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 8),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="regular",
        source_type="generated",
    )
    db.session.add(billable)
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.required_deposit_charge == 0
    assert advice.assessment_deposit_charge == 0
    assert advice.subtotal_sessions == 550.0


def test_billing_total_due_matches_components_without_self_old_balance(session):
    s, t = setup_basic()
    billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 8),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="regular",
        source_type="generated",
    )
    db.session.add(billable)
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    expected = round(
        advice.subtotal_sessions
        + advice.old_balance
        + advice.required_deposit_charge
        + advice.assessment_deposit_charge
        - advice.overpayment_credit,
        2,
    )
    assert advice.total_due == expected

    regenerated = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    expected_regen = round(
        regenerated.subtotal_sessions
        + regenerated.old_balance
        + regenerated.required_deposit_charge
        + regenerated.assessment_deposit_charge
        - regenerated.overpayment_credit,
        2,
    )
    assert regenerated.total_due == expected_regen


def test_legacy_deposit_misclassified_credit_does_not_zero_new_billing(session):
    s, t = setup_basic()
    s.required_deposit_total = 4400
    s.required_deposit_paid = 0
    s.assessment_deposit_total = 500
    s.assessment_deposit_paid = 0
    s.overpayment_credit = 4400  # legacy buggy value
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=4400, purpose="Required Deposit", overpayment_amount=4400))
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, purpose="Assessment Deposit", overpayment_amount=500))

    billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 8),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="regular",
        source_type="generated",
    )
    db.session.add(billable)
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.required_deposit_charge == 0
    assert advice.assessment_deposit_charge == 0
    assert advice.overpayment_credit == 0
    assert advice.subtotal_sessions == 550.0
    assert advice.total_due == 550.0


def test_billing_page_shows_correct_generated_advice_context(session, client):
    s, t = setup_basic()
    s.required_deposit_total = 1000
    s.required_deposit_paid = 1000
    billable = AttendanceSession(
        student_id=s.id,
        therapist_id=t.id,
        session_date=date(2026, 1, 8),
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="Present",
        session_type="regular",
        source_type="generated",
    )
    db.session.add(billable)
    db.session.commit()

    res = client.post(
        "/billing",
        data={"student_id": str(s.id), "start_date": "2026-01-01", "end_date": "2026-01-15"},
        follow_redirects=True,
    )
    assert res.status_code == 200
    assert b"Session Subtotal" in res.data
    assert b"Req: 0" in res.data
    assert b"Total Due" in res.data


def test_billing_total_due_includes_old_unpaid_balance(session):
    s, t = setup_basic()
    c1 = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    c2 = generate_billing_cycles_for_range(date(2026, 1, 16), date(2026, 1, 31))[0]

    prior_advice = BillingAdvice(student_id=s.id, billing_cycle_id=c1.id, total_due=1000, status="Open")
    db.session.add(prior_advice)
    db.session.add(
        AttendanceSession(
            student_id=s.id,
            therapist_id=t.id,
            session_date=date(2026, 1, 20),
            start_time=time(9, 0),
            end_time=time(10, 0),
            duration_hours=1,
            status="Present",
            session_type="regular",
            source_type="generated",
        )
    )
    db.session.commit()

    advice = generate_billing_advices_for_cycle(c2.id, student_id=s.id)[0]
    assert advice.old_balance == 1000
    assert advice.total_due >= advice.subtotal_sessions + 1000


def test_billing_uses_real_carryforward_overpayment_only(session):
    s, t = setup_basic()
    s.overpayment_credit = 300
    db.session.add(
        AttendanceSession(
            student_id=s.id,
            therapist_id=t.id,
            session_date=date(2026, 1, 8),
            start_time=time(9, 0),
            end_time=time(10, 0),
            duration_hours=1,
            status="Present",
            session_type="regular",
            source_type="generated",
        )
    )
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    assert advice.overpayment_credit == 300
    assert advice.total_due == max(
        round(
            advice.subtotal_sessions
            + advice.old_balance
            + advice.required_deposit_charge
            + advice.assessment_deposit_charge
            - 300,
            2,
        ),
        0,
    )


def test_required_deposit_payment_does_not_become_fake_overpayment_when_total_uninitialized(session):
    s, _ = setup_basic()
    # Simulate fresh student before any billing generation initialized required deposit total.
    s.required_deposit_total = 0
    s.required_deposit_paid = 0
    s.overpayment_credit = 0
    db.session.commit()

    payment = record_payment(s.id, 2200, date(2026, 1, 10), purpose="Required Deposit")
    db.session.refresh(s)

    assert s.required_deposit_total == 2200
    assert s.required_deposit_paid == 2200
    assert payment.overpayment_amount == 0
    assert s.overpayment_credit == 0
    assert payment.balance_after_payment == 0


def test_billing_advice_edit_page_loads(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    res = client.get(f"/billing/{advice.id}/edit")
    assert res.status_code == 200
    assert b"Edit Billing Advice" in res.data


def test_editing_billing_advice_updates_record(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    res = client.post(
        f"/billing/{advice.id}/edit",
        data={
            "cycle_start": "2026-01-02",
            "cycle_end": "2026-01-16",
            "issue_date": "2026-01-16",
            "due_date": "2026-01-21",
            "regular_rendered_hours": "1.50",
            "makeup_rendered_hours": "0.50",
            "billed_hours": "0.50",
            "total_rendered_hours": "2.50",
            "billable_hours": "2.00",
            "non_billable_hours": "0.50",
            "session_subtotal": "999.00",
            "required_deposit": "100.00",
            "assessment_deposit": "200.00",
            "old_balance": "300.00",
            "credit": "50.00",
            "total_due": "1549.00",
            "status": "Draft",
            "remarks": "manual correction",
        },
        follow_redirects=True,
    )
    assert res.status_code == 200

    db.session.refresh(advice)
    assert advice.billing_cycle.start_date == date(2026, 1, 2)
    assert advice.billing_cycle.end_date == date(2026, 1, 16)
    assert advice.subtotal_sessions == 999.0
    assert advice.total_due == 1549.0
    assert advice.status == "Draft"


def test_deleting_billing_advice_works_when_safe(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    res = client.post(f"/billing/{advice.id}/delete", follow_redirects=True)
    assert res.status_code == 200
    assert BillingAdvice.query.get(advice.id) is None


def test_deleting_billing_advice_blocked_when_linked_payment_allocations(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    payment = Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100)
    db.session.add(payment)
    db.session.flush()
    db.session.add(PaymentAllocation(payment_id=payment.id, billing_advice_id=advice.id, allocation_type="current_bill", amount=100))
    db.session.commit()

    res = client.post(f"/billing/{advice.id}/delete", follow_redirects=True)
    assert res.status_code == 200
    assert b"Cannot delete billing advice with linked payment allocations" in res.data
    assert BillingAdvice.query.get(advice.id) is not None


def test_billing_page_shows_edit_and_delete_actions(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    generate_billing_advices_for_cycle(cycle.id, student_id=s.id)

    res = client.get(f"/billing?student_id={s.id}")
    assert res.status_code == 200
    assert b"Edit" in res.data
    assert b"Delete" in res.data


def test_attendance_status_billed_is_accepted_and_saved(session, client):
    setup_basic()
    generate_monthly_sessions(2026, 1)
    target = AttendanceSession.query.first()

    res = client.post(
        "/attendance/daily",
        data={
            "selected_date": target.session_date.isoformat(),
            "session_ids": [str(target.id)],
            f"status_{target.id}": "Billed",
        },
        follow_redirects=True,
    )
    assert res.status_code == 200
    db.session.refresh(target)
    assert target.status == "Billed"


def test_billed_counts_as_billable_and_subtotal(session):
    s, t = setup_basic()
    db.session.add(
        AttendanceSession(
            student_id=s.id,
            therapist_id=t.id,
            session_date=date(2026, 1, 3),
            start_time=time(9, 0),
            end_time=time(10, 0),
            duration_hours=1,
            session_type="regular",
            source_type="generated",
            status="Billed",
        )
    )
    db.session.commit()

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    hours = billing_hours_breakdown(s.id, cycle.start_date, cycle.end_date)

    assert hours["billed_hours"] == 1
    assert hours["billable_hours"] == 1
    assert advice.subtotal_sessions == 600


def test_billing_breakdown_and_export_include_billed_hours(session, tmp_path):
    s, t = setup_basic()
    db.session.add(
        AttendanceSession(
            student_id=s.id,
            therapist_id=t.id,
            session_date=date(2026, 1, 2),
            start_time=time(9, 0),
            end_time=time(10, 0),
            duration_hours=1,
            session_type="regular",
            source_type="generated",
            status="Billed",
        )
    )
    db.session.commit()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    generate_billing_advices_for_cycle(cycle.id, student_id=s.id)

    output = export_billing_advices(str(tmp_path / "billing_billed.xlsx"))
    wb = load_workbook(output)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert "Billed Hours" in header


def test_non_billable_logic_still_works_with_billed_status(session):
    s, t = setup_basic()
    db.session.add(
        AttendanceSession(
            student_id=s.id,
            therapist_id=t.id,
            session_date=date(2026, 1, 2),
            start_time=time(9, 0),
            end_time=time(10, 0),
            duration_hours=1,
            session_type="regular",
            source_type="generated",
            status="Non-billable",
        )
    )
    db.session.add(
        AttendanceSession(
            student_id=s.id,
            therapist_id=t.id,
            session_date=date(2026, 1, 3),
            start_time=time(9, 0),
            end_time=time(10, 0),
            duration_hours=1,
            session_type="regular",
            source_type="generated",
            status="Billed",
        )
    )
    db.session.commit()

    hours = billing_hours_breakdown(s.id, date(2026, 1, 1), date(2026, 1, 15))
    assert hours["non_billable_hours"] == 1
    assert hours["billable_hours"] == 1


def test_backup_creates_copy_when_db_exists(tmp_path):
    db_file = tmp_path / "app.db"
    db_file.write_text("db-bytes")
    backup = backup_sqlite_database(f"sqlite:///{db_file}")
    assert backup is not None
    assert backup.exists()
    assert backup.parent.name == "backups"


def test_backup_retention_keeps_latest_seven(tmp_path):
    db_file = tmp_path / "app.db"
    db_file.write_text("db-bytes")
    backup_dir = tmp_path / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)

    for i in range(9):
        p = backup_dir / f"app_20260101_00000{i}.db"
        p.write_text(str(i))

    backup_sqlite_database(f"sqlite:///{db_file}", keep=7)
    assert len(list(backup_dir.glob("app_*.db"))) == 7


def test_backup_creation_with_official_relative_uri_uses_data_backups(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    live_db = tmp_path / "data" / "app.db"
    live_db.parent.mkdir(parents=True, exist_ok=True)
    live_db.write_text("live-db")

    backup = backup_sqlite_database("sqlite:///data/app.db")
    assert backup is not None
    assert backup.parent == tmp_path / "data" / "backups"
    assert backup.name.startswith("app_")


def test_backup_listing_with_official_relative_uri_reads_data_backups(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    live_db = tmp_path / "data" / "app.db"
    backup_dir = tmp_path / "data" / "backups"
    live_db.parent.mkdir(parents=True, exist_ok=True)
    backup_dir.mkdir(parents=True, exist_ok=True)
    live_db.write_text("live-db")
    (backup_dir / "app_20260101_000001.db").write_text("b1")

    listed = list_sqlite_backups("sqlite:///data/app.db")
    assert listed
    assert listed[0].parent == backup_dir


def test_backup_listing_uses_database_sibling_backups_directory(tmp_path):
    db_file = tmp_path / "nested" / "db" / "app.db"
    db_file.parent.mkdir(parents=True, exist_ok=True)
    db_file.write_text("db-bytes")
    backup_path = backup_sqlite_database(f"sqlite:///{db_file}")
    assert backup_path is not None

    listed = list_sqlite_backups(f"sqlite:///{db_file}")
    assert listed
    assert listed[0].parent == db_file.parent / "backups"
    assert listed[0].name.startswith("app_")
    assert listed[0].suffix == ".db"


def test_official_sqlite_relative_uri_resolves_to_project_data_app_db(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    resolved = resolve_sqlite_db_path("sqlite:///data/app.db")
    assert resolved == tmp_path / "data" / "app.db"


def test_legacy_instance_database_migrates_when_official_missing(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    legacy_path = tmp_path / "instance" / "data" / "app.db"
    legacy_path.parent.mkdir(parents=True, exist_ok=True)
    legacy_path.write_text("legacy-db")

    app = create_app({"TESTING": True, "SQLALCHEMY_DATABASE_URI": "sqlite:///data/app.db"})
    official_path = tmp_path / "data" / "app.db"
    assert app.config["SQLALCHEMY_DATABASE_URI"] == f"sqlite:///{official_path}"
    assert official_path.read_text() == "legacy-db"


def test_legacy_migration_does_not_overwrite_existing_official_db(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    legacy_path = tmp_path / "instance" / "data" / "app.db"
    legacy_path.parent.mkdir(parents=True, exist_ok=True)
    legacy_path.write_text("legacy-db")
    official_path = tmp_path / "data" / "app.db"
    official_path.parent.mkdir(parents=True, exist_ok=True)
    official_path.write_text("official-db")

    create_app({"TESTING": True, "SQLALCHEMY_DATABASE_URI": "sqlite:///data/app.db"})
    assert official_path.read_text() == "official-db"


def test_sqlite_parent_directory_is_auto_created_when_missing(tmp_path, monkeypatch):
    from app import _ensure_sqlite_parent_directory
    from flask import Flask

    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    app = Flask(__name__, instance_relative_config=True)
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///data/app.db"
    target_dir = tmp_path / "data"
    assert not target_dir.exists()
    _ensure_sqlite_parent_directory(app)

    assert target_dir.exists()


def test_init_db_no_longer_requires_manual_instance_data_creation(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    app = create_app({"TESTING": True, "SQLALCHEMY_DATABASE_URI": "sqlite:///data/app.db"})
    target_dir = tmp_path / "data"
    db_file = target_dir / "app.db"

    result = app.test_cli_runner().invoke(args=["init-db"])

    assert result.exit_code == 0
    assert target_dir.exists()
    assert db_file.exists()


def test_existing_sqlite_parent_directory_remains_safe(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    seed_app = create_app({"TESTING": True, "SQLALCHEMY_DATABASE_URI": "sqlite:///data/app.db"})
    target_dir = tmp_path / "data"
    target_dir.mkdir(parents=True, exist_ok=True)
    app = create_app({"TESTING": True, "SQLALCHEMY_DATABASE_URI": "sqlite:///data/app.db"})
    result = app.test_cli_runner().invoke(args=["init-db"])

    assert result.exit_code == 0
    assert target_dir.exists()


def test_non_sqlite_database_uri_does_not_create_sqlite_directories(tmp_path, monkeypatch):
    from app import _ensure_sqlite_parent_directory
    from flask import Flask

    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    app = Flask(__name__, instance_relative_config=True)
    app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://user:pw@localhost/db"
    _ensure_sqlite_parent_directory(app)

    assert not (tmp_path / "data").exists()


def test_system_no_longer_uses_instance_data_as_live_db_after_migration(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.backup_utils.project_root_path", lambda: tmp_path)
    legacy_path = tmp_path / "instance" / "data" / "app.db"
    legacy_path.parent.mkdir(parents=True, exist_ok=True)
    legacy_path.write_text("legacy-db")

    app = create_app({"TESTING": True, "SQLALCHEMY_DATABASE_URI": "sqlite:///data/app.db"})
    official_path = tmp_path / "data" / "app.db"

    assert app.config["SQLALCHEMY_DATABASE_URI"] == f"sqlite:///{official_path}"
    assert official_path.exists()


def test_editing_issued_billing_advice_is_blocked(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    advice.status = "Issued"
    db.session.commit()

    res = client.post(
        f"/billing/{advice.id}/edit",
        data={
            "cycle_start": "2026-01-02",
            "cycle_end": "2026-01-16",
            "issue_date": "2026-01-16",
            "due_date": "2026-01-21",
            "session_subtotal": "999.00",
        },
        follow_redirects=True,
    )
    assert res.status_code == 200
    assert b"Frozen Billing Correction Approval" in res.data


def test_deleting_issued_paid_archived_billing_advice_is_blocked(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]

    for status in ["Issued", "Paid", "Archived"]:
        advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
        advice.status = status
        db.session.commit()

        res = client.post(f"/billing/{advice.id}/delete", follow_redirects=True)
        assert res.status_code == 200
        assert b"locked and cannot be deleted" in res.data


def test_marking_billing_advice_status_updates(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    res = client.post(f"/billing/{advice.id}/status", data={"status": "Issued"}, follow_redirects=True)
    assert res.status_code == 200
    db.session.refresh(advice)
    assert advice.status == "Issued"


def test_audit_logs_created_for_key_actions(session, client):
    s, _ = setup_basic()
    generate_monthly_sessions(2026, 1)
    target = AttendanceSession.query.first()

    client.post(
        "/attendance/daily",
        data={"selected_date": target.session_date.isoformat(), "session_ids": [str(target.id)], f"status_{target.id}": "Present"},
        follow_redirects=True,
    )

    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    client.post(
        f"/billing/{advice.id}/edit",
        data={
            "cycle_start": cycle.start_date.isoformat(),
            "cycle_end": cycle.end_date.isoformat(),
            "issue_date": cycle.issue_date.isoformat(),
            "due_date": cycle.due_date.isoformat(),
            "session_subtotal": "777",
        },
        follow_redirects=True,
    )
    client.post(f"/billing/{advice.id}/status", data={"status": "Issued"}, follow_redirects=True)
    client.post(f"/billing/{advice.id}/status", data={"status": "Paid"}, follow_redirects=True)
    client.post(f"/billing/{advice.id}/status", data={"status": "Archived"}, follow_redirects=True)

    c2 = generate_billing_cycles_for_range(date(2026, 1, 16), date(2026, 1, 31))[0]
    advice_to_delete = generate_billing_advices_for_cycle(c2.id, student_id=s.id)[0]
    client.post(f"/billing/{advice_to_delete.id}/delete", follow_redirects=True)

    client.post(f"/payments", data={"student_id": s.id, "amount": 100, "payment_date": "2026-01-10"}, follow_redirects=True)
    payment = Payment.query.order_by(Payment.id.desc()).first()
    client.post(
        f"/payments/{payment.id}/edit",
        data={
            "payment_date": "2026-01-10",
            "student_id": s.id,
            "purpose": "Therapy",
            "amount": 100,
            "overpayment_amount": 0,
            "balance_after_payment": 0,
        },
        follow_redirects=True,
    )

    actions = {a.action for a in AuditLog.query.all()}
    assert "billing_advice_created" in actions
    assert "billing_advice_edited" in actions
    assert "billing_advice_marked_issued" in actions
    assert "billing_advice_marked_paid" in actions
    assert "billing_advice_marked_archived" in actions
    assert "billing_advice_deleted" in actions
    assert "payment_recorded" in actions
    assert "payment_edited" in actions
    assert "attendance_status_updated" in actions


def test_audit_logging_failure_does_not_crash_workflow(session, monkeypatch):
    s, _ = setup_basic()
    import app.services.payment_service as payment_service

    def boom(*args, **kwargs):
        raise RuntimeError("audit failed")

    monkeypatch.setattr(payment_service, "log_audit", boom)
    payment = record_payment(s.id, 100, date(2026, 1, 10))
    assert payment.id is not None


def test_restore_replaces_database_from_selected_backup(tmp_path):
    db_file = tmp_path / "app.db"
    db_file.write_text("live-db")
    backup_dir = tmp_path / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_file = backup_dir / "app_20260101_000000.db"
    backup_file.write_text("backup-db")

    restore_sqlite_backup(f"sqlite:///{db_file}", backup_file)

    assert db_file.read_text() == "backup-db"


def test_restore_route_rejects_invalid_backup_filename(client):
    res = client.post("/restore-backup/not_a_backup.txt", follow_redirects=True)

    assert res.status_code == 200
    assert b"Invalid backup filename." in res.data


def test_restore_route_restores_selected_backup(app, client, tmp_path):
    db_file = tmp_path / "app.db"
    db_file.write_text("live-db")
    backup_dir = tmp_path / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_file = backup_dir / "app_20260101_000000.db"
    backup_file.write_text("backup-db")

    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_file}"

    res = client.post(f"/restore-backup/{backup_file.name}", follow_redirects=True)

    assert res.status_code == 200
    assert b"Backup restored." in res.data
    assert db_file.read_text() == "backup-db"


def test_create_backup_route_creates_file_in_instance_data_backups(app, client, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    db_file = tmp_path / "instance" / "data" / "app.db"
    db_file.parent.mkdir(parents=True, exist_ok=True)
    db_file.write_text("live-db")

    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///instance/data/app.db"
    res = client.post("/create-backup", follow_redirects=True)

    assert res.status_code == 200
    assert b"Backup created:" in res.data
    backups = sorted((tmp_path / "instance" / "data" / "backups").glob("app_*.db"))
    assert len(backups) == 1


def test_dashboard_lists_backups_after_manual_creation(app, client, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    db_file = tmp_path / "instance" / "data" / "app.db"
    db_file.parent.mkdir(parents=True, exist_ok=True)
    db_file.write_text("live-db")
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///instance/data/app.db"

    client.post("/create-backup", follow_redirects=True)
    page = client.get("/")
    backup_names = [p.name.encode() for p in (tmp_path / "instance" / "data" / "backups").glob("app_*.db")]
    assert backup_names
    assert any(name in page.data for name in backup_names)


def test_create_backup_route_rejects_non_sqlite_database(app, client):
    app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://user:pw@localhost/db"

    res = client.post("/create-backup", follow_redirects=True)

    assert res.status_code == 200
    assert b"Backup is only supported for sqlite file databases." in res.data


class _FixedDate(date):
    @classmethod
    def today(cls):
        return cls(2026, 3, 31)


class _SundayDate(date):
    @classmethod
    def today(cls):
        return cls(2026, 1, 4)


class _AprilDate(date):
    @classmethod
    def today(cls):
        return cls(2026, 4, 1)


class _MondayDate(date):
    @classmethod
    def today(cls):
        return cls(2026, 1, 5)


def test_dashboard_billings_today_card_shows_students_due_for_generation(session, client):
    s, _ = setup_basic()
    s.created_at = datetime.combine(date.today() - timedelta(days=14), datetime.min.time())
    db.session.commit()

    res = client.get('/')
    assert res.status_code == 200
    assert b'Billings Today' in res.data
    assert bytes(s.name, 'utf-8') in res.data


def test_monthly_archive_reminder_pending_after_month_end_if_prior_month_missing(session, client, monkeypatch):
    monkeypatch.setattr('app.routes.web.date', _AprilDate)
    res = client.get('/')
    assert res.status_code == 200
    assert b'Monthly Ledger Archive' in res.data
    assert b'Archive for March 2026 is still pending.' in res.data


def test_monthly_archive_reminder_completed_after_month_end_when_prior_month_archived(session, client, monkeypatch):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 3, 10), amount=100, is_archived=True, archive_month=3, archive_year=2026))
    db.session.commit()

    monkeypatch.setattr('app.routes.web.date', _AprilDate)
    res = client.get('/')
    assert res.status_code == 200
    assert b'Monthly archive is complete for March 2026.' in res.data


def test_monthly_archive_duplicate_prevented(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500))
    db.session.commit()

    first = client.post('/payments/tracker', data={'action': 'archive', 'archive_month': '1', 'archive_year': '2026'}, follow_redirects=True)
    second = client.post('/payments/tracker', data={'action': 'archive', 'archive_month': '1', 'archive_year': '2026'}, follow_redirects=True)
    assert first.status_code == 200
    assert b'January 2026 ledger archived successfully (1 record(s)).' in first.data
    assert b'Archive for 2026-01 already exists.' in second.data


def test_archived_ledger_month_year_list_and_open(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 3, 11), amount=100, is_archived=True, archive_month=3, archive_year=2026))
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 2, 8), amount=120, is_archived=True, archive_month=2, archive_year=2026))
    db.session.commit()

    res = client.get('/payments/tracker?view=archived')
    assert res.status_code == 200
    assert b'Archived Month-Year' in res.data
    assert b'March 2026' in res.data
    detail = client.get('/payments/tracker?view=archived&month=2&year=2026')
    assert b'2026-02-08' in detail.data


def test_payment_ledger_search_matches_text_fields(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, client_guardian_name='Guardian Match', purpose='Therapy Booster', billing_period_start=date(2026, 1, 1), billing_period_end=date(2026, 1, 15)))
    db.session.commit()

    assert b'2026-01-10' in client.get('/payments/tracker?view=active&month=1&year=2026&q=S').data
    assert b'2026-01-10' in client.get('/payments/tracker?view=active&month=1&year=2026&q=guardian').data
    assert b'2026-01-10' in client.get('/payments/tracker?view=active&month=1&year=2026&q=2026-01-01').data
    assert b'2026-01-10' in client.get('/payments/tracker?view=active&month=1&year=2026&q=booster').data


def test_empty_payment_search_shows_default_listing(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, client_guardian_name='A'))
    db.session.commit()

    res = client.get('/payments/tracker?view=active&month=1&year=2026&q=')
    assert res.status_code == 200
    assert b'2026-01-10' in res.data


def test_weekly_archive_reminder_pending_after_sunday_if_week_missing(session, client, monkeypatch):
    monkeypatch.setattr('app.routes.web.date', _MondayDate)
    res = client.get('/')
    assert res.status_code == 200
    assert b'Weekly Report Archive' in res.data
    assert b'Archive for 2025-12-29 to 2026-01-04 is still pending.' in res.data


def test_weekly_archive_reminder_completed_after_sunday_when_week_archived(session, client, monkeypatch):
    monkeypatch.setattr('app.routes.web.date', _MondayDate)
    db.session.add(WeeklyReportArchive(week_start=date(2025, 12, 29), week_end=date(2026, 1, 4), note='done'))
    db.session.commit()

    res = client.get('/')
    assert res.status_code == 200
    assert b'No weekly archive action needed right now.' in res.data


def test_navigation_order_and_tabs(session, client):
    res = client.get('/')
    assert res.status_code == 200
    html = res.data.decode('utf-8')
    assert html.index('>Dashboard<') < html.index('>Daily Schedule<') < html.index('>Make-up Editor<')
    assert 'Import' not in html
    assert 'Admin Attendance</a>' not in html


def test_admin_attendance_is_available_in_dashboard(session, client):
    admin = AdminStaff(name='Dashboard Admin')
    db.session.add(admin)
    db.session.commit()

    res = client.post('/', data={
        'action': 'save_admin_attendance',
        'admin_id': str(admin.id),
        'attendance_date': '2026-01-10',
        'status': 'Present',
        'shift_label': 'AM',
        'hours_worked': '4',
    }, follow_redirects=True)
    assert res.status_code == 200
    assert b'Admin attendance saved.' in res.data
    assert b'Admin Attendance (Dashboard)' in res.data


def test_dashboard_renders_grouped_sections_and_priority_open(session, client):
    res = client.get('/')
    assert res.status_code == 200
    html = res.data.decode('utf-8')
    assert 'Priority Today' in html
    assert 'Daily Operations' in html
    assert 'System / Utilities' in html
    assert '<details class="dashboard-group" id="priority-today" open>' in html


def test_dashboard_lower_priority_sections_collapsible(session, client):
    res = client.get('/')
    assert res.status_code == 200
    html = res.data.decode('utf-8')
    assert html.count('<details class="dashboard-group"') >= 3


def test_dashboard_renders_today_priorities_summary_strip(session, client):
    res = client.get('/')
    assert res.status_code == 200
    assert b"Today's priorities:" in res.data


def test_dashboard_action_cards_show_status_and_primary_links(session, client):
    res = client.get('/')
    assert res.status_code == 200
    assert b'Status:' in res.data
    assert b'Go to Billing' in res.data
    assert b'Go to Archive' in res.data
    assert b'Go to Weekly Archive' in res.data


def test_dashboard_empty_states_and_standardized_status_labels(session, client):
    res = client.get('/')
    assert res.status_code == 200
    assert b'No billing tasks need action today.' in res.data
    assert b'Status: No action needed' in res.data


def test_dashboard_shows_last_completed_archive_context_or_fallback(session, client):
    res = client.get('/')
    assert res.status_code == 200
    assert b'Last completed: No archive recorded yet.' in res.data


def test_dashboard_quick_links_block_renders_expected_destinations(session, client):
    res = client.get('/')
    assert res.status_code == 200
    assert b'Quick Links' in res.data
    assert b'Daily Schedule' in res.data
    assert b'Make-up Editor' in res.data
    assert b'Billing' in res.data
    assert b'Payment Ledger' in res.data


def test_payment_ledger_search_ui_shows_summary_and_clear_link(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=500, client_guardian_name='Search Person'))
    db.session.commit()

    res = client.get('/payments/tracker?view=active&month=1&year=2026&q=Search')
    assert res.status_code == 200
    assert b'Showing 1 matching payment entry for "Search".' in res.data
    assert b'Clear search' in res.data


def test_admin_attendance_save_message_is_specific(session, client):
    admin = AdminStaff(name='Message Admin')
    db.session.add(admin)
    db.session.commit()

    res = client.post('/', data={
        'action': 'save_admin_attendance',
        'admin_id': str(admin.id),
        'attendance_date': '2026-01-10',
        'status': 'Present',
        'shift_label': 'AM',
        'hours_worked': '4',
    }, follow_redirects=True)
    assert res.status_code == 200
    assert b'Admin attendance saved.' in res.data


def test_student_required_deposit_policy_flag_defaults_true(session):
    s = Student(name="Policy Default", contract_hours_per_week=2)
    db.session.add(s)
    db.session.commit()

    assert s.required_deposit_enabled is True


def test_student_can_be_marked_no_required_deposit(session, client):
    res = client.post('/master-data/students', data={
        'action': 'create',
        'name': 'No Deposit Student',
        'contract_hours_per_week': '3',
        'required_deposit_enabled': '0',
        'overpayment_credit': '0',
        'active': '1',
    }, follow_redirects=True)
    assert res.status_code == 200
    student = Student.query.filter_by(name='No Deposit Student').first()
    assert student is not None
    assert student.required_deposit_enabled is False


def test_billing_skips_required_deposit_when_disabled(session):
    s, _ = setup_basic()
    s.required_deposit_enabled = False
    db.session.commit()

    generate_monthly_sessions(2026, 1)
    mark_rendered(s.id)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]

    assert advice.required_deposit_charge == 0


def test_required_deposit_ui_shows_policy_options(session, client):
    res = client.get('/master-data/students')
    assert res.status_code == 200
    assert b'With Required Deposit' in res.data
    assert b'No Required Deposit' in res.data


def test_payment_create_in_archived_month_marks_archive_outdated(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 5), amount=100))
    db.session.commit()
    client.post('/payments/tracker', data={'action': 'archive', 'archive_month': '1', 'archive_year': '2026'}, follow_redirects=True)

    client.post('/payments', data={
        'payment_date': '2026-01-10', 'client_guardian_name': 'A', 'student_id': str(s.id), 'purpose': 'Therapy', 'amount': '50'
    }, follow_redirects=True)

    snapshot = MonthlyPaymentArchive.query.filter_by(archive_month=1, archive_year=2026).first()
    assert snapshot is not None
    assert snapshot.status == 'outdated'


def test_payment_edit_in_archived_month_marks_archive_outdated(session, client):
    s, _ = setup_basic()
    p = Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100)
    db.session.add(p)
    db.session.commit()
    client.post('/payments/tracker', data={'action': 'archive', 'archive_month': '1', 'archive_year': '2026'}, follow_redirects=True)

    res = client.post(f'/payments/{p.id}/edit', data={'payment_date': '2026-01-11', 'client_guardian_name': 'X', 'student_id': str(s.id), 'purpose': 'Therapy', 'amount': '200', 'mode_of_transfer': 'Cash'}, follow_redirects=True)
    assert res.status_code == 200
    snapshot = MonthlyPaymentArchive.query.filter_by(archive_month=1, archive_year=2026).first()
    assert snapshot.status == 'outdated'


def test_payment_delete_in_archived_month_marks_archive_outdated(session, client):
    s, _ = setup_basic()
    p = Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100)
    db.session.add(p)
    db.session.commit()
    client.post('/payments/tracker', data={'action': 'archive', 'archive_month': '1', 'archive_year': '2026'}, follow_redirects=True)

    res = client.post(f'/payments/{p.id}/delete', follow_redirects=True)
    assert res.status_code == 200
    snapshot = MonthlyPaymentArchive.query.filter_by(archive_month=1, archive_year=2026).first()
    assert snapshot.status == 'outdated'


def test_archived_list_shows_outdated_status_and_refresh_link(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100, is_archived=True, archive_month=1, archive_year=2026))
    db.session.add(MonthlyPaymentArchive(archive_month=1, archive_year=2026, status='outdated', archived_total_amount=100, archived_entry_count=1))
    db.session.commit()

    res = client.get('/payments/tracker?view=archived&month=1&year=2026')
    assert b'Status: Outdated' in res.data
    assert b'Refresh Archive' in res.data


def test_supervisor_master_create_hashes_password(session, client):
    res = client.post('/master-data/supervisors', data={'action': 'create', 'name': 'Sup A', 'role': 'Billing Lead', 'password': 'secret123'}, follow_redirects=True)
    assert res.status_code == 200
    row = Supervisor.query.filter_by(name='Sup A').first()
    assert row is not None
    assert row.password_hash != 'secret123'


def test_supervisor_password_not_exposed_in_ui(session, client):
    client.post('/master-data/supervisors', data={'action': 'create', 'name': 'Sup B', 'role': 'Billing Lead', 'password': 'secret123'}, follow_redirects=True)
    res = client.get('/master-data/supervisors')
    assert b'secret123' not in res.data


def test_inactive_supervisor_cannot_approve_refresh(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100, is_archived=True, archive_month=1, archive_year=2026))
    db.session.add(MonthlyPaymentArchive(archive_month=1, archive_year=2026, status='outdated', archived_total_amount=100, archived_entry_count=1))
    from werkzeug.security import generate_password_hash
    db.session.add(Supervisor(name='Inactive Sup', role='Admin Supervisor', password_hash=generate_password_hash('pw'), is_active=False))
    db.session.commit()

    res = client.post('/payments/archive-review/2026/1', data={'supervisor_name': 'Inactive Sup', 'supervisor_password': 'pw', 'refresh_reason': 'test'}, follow_redirects=True)
    assert b'Supervisor approval failed' in res.data


def test_review_page_shows_month_status_and_comparison(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100))
    db.session.add(MonthlyPaymentArchive(archive_month=1, archive_year=2026, status='outdated', archived_total_amount=80, archived_entry_count=1))
    db.session.commit()

    res = client.get('/payments/archive-review/2026/1')
    assert b'Archive Refresh Review - January 2026' in res.data
    assert b'Archive status: <strong>Outdated</strong>' in res.data
    assert b'Old archived total amount' in res.data
    assert b'New recalculated total amount' in res.data
    assert b'Supervisor Approval' in res.data


def test_wrong_supervisor_password_blocks_refresh(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100, is_archived=True, archive_month=1, archive_year=2026))
    db.session.add(MonthlyPaymentArchive(archive_month=1, archive_year=2026, status='outdated', archived_total_amount=100, archived_entry_count=1))
    from werkzeug.security import generate_password_hash
    db.session.add(Supervisor(name='Sup C', role='Admin Supervisor', password_hash=generate_password_hash('goodpw'), is_active=True))
    db.session.commit()

    res = client.post('/payments/archive-review/2026/1', data={'supervisor_name': 'Sup C', 'supervisor_password': 'badpw', 'refresh_reason': 'fix'}, follow_redirects=True)
    assert b'Supervisor approval failed' in res.data


def test_correct_supervisor_refreshes_archive_and_logs_audit(session, client):
    s, _ = setup_basic()
    db.session.add(Payment(student_id=s.id, payment_date=date(2026, 1, 10), amount=100, is_archived=True, archive_month=1, archive_year=2026))
    db.session.add(MonthlyPaymentArchive(archive_month=1, archive_year=2026, status='outdated', archived_total_amount=50, archived_entry_count=1))
    from werkzeug.security import generate_password_hash
    db.session.add(Supervisor(name='Sup D', role='Admin Supervisor', password_hash=generate_password_hash('goodpw'), is_active=True))
    db.session.commit()

    res = client.post('/payments/archive-review/2026/1', data={'supervisor_name': 'Sup D', 'supervisor_password': 'goodpw', 'refresh_reason': 'month close'}, follow_redirects=True)
    assert b'January 2026 archive refreshed successfully.' in res.data

    snapshot = MonthlyPaymentArchive.query.filter_by(archive_month=1, archive_year=2026).first()
    assert snapshot.status == 'current'
    audit = AuditLog.query.filter_by(action='archive_refresh_approved').first()
    assert audit is not None
    assert '2026-01' in audit.details


def test_generating_billing_removes_student_from_billings_today(session, client, monkeypatch):
    s, _ = setup_basic()
    s.created_at = datetime.combine(date(2026, 1, 1), datetime.min.time())
    db.session.commit()

    import app.routes.web as web_routes
    class _BillDate(date):
        @classmethod
        def today(cls):
            return cls(2026, 1, 15)
    monkeypatch.setattr(web_routes, "date", _BillDate)

    before = client.get('/')
    assert b'(2026-01-01 to 2026-01-15)' in before.data

    client.post('/billing', data={'student_id': str(s.id), 'start_date': '2026-01-01', 'end_date': '2026-01-15'}, follow_redirects=True)
    after = client.get('/')
    assert b'(2026-01-01 to 2026-01-15)' not in after.data


def test_upcoming_and_overdue_use_unsettled_generated_billing(session, client):
    s, _ = setup_basic()
    c1 = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    a1 = generate_billing_advices_for_cycle(c1.id, student_id=s.id)[0]
    a1.billing_cycle.due_date = date.today() + timedelta(days=2)

    c2 = generate_billing_cycles_for_range(date(2026, 1, 16), date(2026, 1, 30))[0]
    a2 = generate_billing_advices_for_cycle(c2.id, student_id=s.id)[0]
    a2.billing_cycle.due_date = date.today() - timedelta(days=2)
    db.session.commit()

    res = client.get('/')
    assert b'Upcoming Dues' in res.data
    assert b'Overdue' in res.data


def test_full_payment_removes_due_overdue_but_partial_does_not(session, client):
    s, _ = setup_basic()
    c = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(c.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    db.session.commit()

    # partial payment keeps overdue
    record_payment(s.id, 100, date.today())
    partial = client.get('/')
    assert b'Overdue' in partial.data

    # full settlement removes overdue
    record_payment(s.id, max(advice.total_due, 0), date.today())
    settled = client.get('/')
    assert b'No overdue billing advice.' in settled.data


def test_frozen_billing_requires_supervisor_approval_for_edit(session, client):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    advice.status = 'Issued'
    from werkzeug.security import generate_password_hash
    db.session.add(Supervisor(name='Bill Sup', role='Billing Lead', password_hash=generate_password_hash('pw'), is_active=True))
    db.session.commit()

    res = client.get(f'/billing/{advice.id}/edit')
    assert b'Frozen Billing Correction Approval' in res.data


def test_correct_supervisor_allows_frozen_billing_edit_and_logs(session, client):
    s, _ = setup_basic()
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    advice.status = 'Issued'
    from werkzeug.security import generate_password_hash
    db.session.add(Supervisor(name='Bill Sup2', role='Billing Lead', password_hash=generate_password_hash('pw'), is_active=True))
    db.session.commit()

    client.post(f'/billing/{advice.id}/approve-frozen-edit', data={'supervisor_name': 'Bill Sup2', 'supervisor_password': 'pw', 'approval_reason': 'client correction'}, follow_redirects=True)
    res = client.post(f'/billing/{advice.id}/edit', data={
        'cycle_start': '2026-01-01', 'cycle_end': '2026-01-15', 'issue_date': '2026-01-15', 'due_date': '2026-01-20',
        'session_subtotal': str(advice.subtotal_sessions), 'required_deposit': str(advice.required_deposit_charge),
        'assessment_deposit': str(advice.assessment_deposit_charge), 'old_balance': str(advice.old_balance),
        'credit': str(advice.overpayment_credit), 'total_due': str(advice.total_due), 'status': 'Issued'
    }, follow_redirects=True)
    assert b'Billing advice updated.' in res.data
    audit = AuditLog.query.filter_by(action='frozen_billing_correction_approved').first()
    assert audit is not None


def test_red_bills_to_issue_shows_overdue_unsettled_without_notice(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    db.session.add(AttendanceSession(student_id=s.id, therapist_id=s.assigned_therapist_id, session_date=date.today()+timedelta(days=1), start_time=time(9,0), end_time=time(10,0), duration_hours=1, status='Present', source_type='manual'))
    db.session.commit()

    res = client.get('/')
    assert b'Red Bills To Issue' in res.data
    assert b'Red Bill Needed' in res.data


def test_issue_red_bill_moves_case_to_active(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    db.session.add(AttendanceSession(student_id=s.id, therapist_id=s.assigned_therapist_id, session_date=date.today()+timedelta(days=2), start_time=time(9,0), end_time=time(10,0), duration_hours=1, status='Present', source_type='manual'))
    db.session.commit()

    client.post(f'/billing/{advice.id}/issue-red-bill', follow_redirects=True)
    res = client.get('/')
    assert b'Red Bills Active' in res.data


def test_suspension_required_shows_when_red_bill_due_passed(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=10)
    db.session.add(RedBillingNotice(billing_advice_id=advice.id, student_id=s.id, issued_date=date.today()-timedelta(days=9), outstanding_amount=advice.total_due, next_session_date=date.today()-timedelta(days=1), red_bill_due_date=date.today()-timedelta(days=2), status='issued'))
    db.session.commit()

    res = client.get('/')
    assert b'Suspension Required' in res.data


def test_red_bill_reminders_removed_when_settled(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    db.session.add(RedBillingNotice(billing_advice_id=advice.id, student_id=s.id, issued_date=date.today(), outstanding_amount=advice.total_due, next_session_date=date.today()+timedelta(days=1), red_bill_due_date=date.today(), status='issued'))
    db.session.commit()

    record_payment(s.id, advice.total_due, date.today())
    res = client.get('/')
    assert b'No active Red Bills at this time.' in res.data


def test_red_bill_due_date_uses_next_session_strictly_after_today(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    db.session.add(AttendanceSession(student_id=s.id, therapist_id=s.assigned_therapist_id, session_date=date.today(), start_time=time(9, 0), end_time=time(10, 0), duration_hours=1, status='Present', source_type='manual'))
    db.session.add(AttendanceSession(student_id=s.id, therapist_id=s.assigned_therapist_id, session_date=date.today() + timedelta(days=3), start_time=time(9, 0), end_time=time(10, 0), duration_hours=1, status='Present', source_type='manual'))
    db.session.commit()

    client.post(f'/billing/{advice.id}/issue-red-bill', follow_redirects=True)
    notice = RedBillingNotice.query.filter_by(billing_advice_id=advice.id).first()
    assert notice.next_session_date == date.today() + timedelta(days=3)
    assert notice.red_bill_due_date == date.today() + timedelta(days=2)


def test_red_bill_no_future_session_is_due_today_immediate(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    db.session.commit()

    res = client.post(f'/billing/{advice.id}/issue-red-bill', follow_redirects=True)
    assert b'Red Bill due date' in res.data
    notice = RedBillingNotice.query.filter_by(billing_advice_id=advice.id).first()
    assert notice.next_session_date is None
    assert notice.red_bill_due_date == date.today()


def test_red_bill_schedule_changes_before_but_not_after_issuance(session, client):
    s, _ = setup_basic()
    cyc = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cyc.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=1)
    original = AttendanceSession(student_id=s.id, therapist_id=s.assigned_therapist_id, session_date=date.today() + timedelta(days=4), start_time=time(9, 0), end_time=time(10, 0), duration_hours=1, status='Present', source_type='manual')
    db.session.add(original)
    db.session.commit()

    original.session_date = date.today() + timedelta(days=2)
    db.session.commit()
    client.post(f'/billing/{advice.id}/issue-red-bill', follow_redirects=True)
    notice = RedBillingNotice.query.filter_by(billing_advice_id=advice.id).first()
    due_at_issue = notice.red_bill_due_date
    assert due_at_issue == date.today() + timedelta(days=1)

    original.session_date = date.today() + timedelta(days=8)
    db.session.commit()
    refreshed = RedBillingNotice.query.filter_by(billing_advice_id=advice.id).first()
    assert refreshed.red_bill_due_date == due_at_issue


def test_suspended_student_daily_attendance_is_locked(session, client):
    s, _ = setup_basic()
    target_date = date.today()
    sess = AttendanceSession(
        student_id=s.id,
        therapist_id=s.assigned_therapist_id,
        session_date=target_date,
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="",
        source_type="manual",
    )
    db.session.add(sess)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=10)
    db.session.add(RedBillingNotice(billing_advice_id=advice.id, student_id=s.id, issued_date=date.today()-timedelta(days=9), outstanding_amount=advice.total_due, next_session_date=date.today()-timedelta(days=1), red_bill_due_date=date.today()-timedelta(days=2), status='issued'))
    db.session.commit()

    page = client.get(f'/attendance/daily?date={target_date.isoformat()}')
    assert b'Suspended' in page.data

    client.post('/attendance/daily', data={'selected_date': target_date.isoformat(), 'session_ids': [str(sess.id)], f'status_{sess.id}': 'Present'}, follow_redirects=True)
    db.session.refresh(sess)
    assert sess.status == 'Suspended'


def test_full_settlement_lifts_suspension_and_partial_needs_exception(session, client):
    s, _ = setup_basic()
    target_date = date.today()
    sess = AttendanceSession(
        student_id=s.id,
        therapist_id=s.assigned_therapist_id,
        session_date=target_date,
        start_time=time(9, 0),
        end_time=time(10, 0),
        duration_hours=1,
        status="",
        source_type="manual",
    )
    db.session.add(sess)
    cycle = generate_billing_cycles_for_range(date(2026, 1, 1), date(2026, 1, 15))[0]
    advice = generate_billing_advices_for_cycle(cycle.id, student_id=s.id)[0]
    advice.billing_cycle.due_date = date.today() - timedelta(days=10)
    db.session.add(RedBillingNotice(billing_advice_id=advice.id, student_id=s.id, issued_date=date.today()-timedelta(days=9), outstanding_amount=advice.total_due, next_session_date=date.today()-timedelta(days=1), red_bill_due_date=date.today()-timedelta(days=2), status='issued'))
    from werkzeug.security import generate_password_hash
    db.session.add(Supervisor(name='Ops Sup', role='Ops', password_hash=generate_password_hash('pw'), is_active=True))
    db.session.commit()

    record_payment(s.id, 100, date.today())
    client.post('/attendance/daily', data={'selected_date': target_date.isoformat(), 'session_ids': [str(sess.id)], f'status_{sess.id}': 'Present'}, follow_redirects=True)
    db.session.refresh(sess)
    assert sess.status == 'Suspended'

    client.post(f'/billing/{advice.id}/suspension-exception', data={'supervisor_name': 'Ops Sup', 'supervisor_password': 'pw', 'override_reason': 'Allow return'}, follow_redirects=True)
    client.post('/attendance/daily', data={'selected_date': target_date.isoformat(), 'session_ids': [str(sess.id)], f'status_{sess.id}': 'Present'}, follow_redirects=True)
    db.session.refresh(sess)
    assert sess.status == 'Present'
    audit = AuditLog.query.filter_by(action='suspension_partial_settlement_lift_approved').first()
    assert audit is not None

    record_payment(s.id, advice.total_due, date.today())
    client.post('/attendance/daily', data={'selected_date': target_date.isoformat(), 'session_ids': [str(sess.id)], f'status_{sess.id}': 'Absent'}, follow_redirects=True)
    db.session.refresh(sess)
    assert sess.status == 'Absent'
